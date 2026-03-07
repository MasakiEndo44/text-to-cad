import argparse
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Please install it using: pip install openpyxl")
    sys.exit(1)

def export_xlsx(json_path, out_path):
    if not os.path.exists(json_path):
        print(f"Error: Could not find {json_path}")
        sys.exit(1)
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    bom_items = data.get("bom", [])
    if not bom_items:
        print("Warning: No BOM data found in the JSON file. Ensure the 'bom' array exists.")
        
    wb = Workbook()
    ws = wb.active
    ws.title = "部品表 (BOM)"
    
    headers = [
        "カテゴリ", "部品番号", "品名", "数量", 
        "用途 / 役割", "材質 / 仕様", "推奨型番\n(ミスミ/モノタロウ等)", "備考 / リンク"
    ]
    
    ws.append(headers)
    
    # -----------------------------
    # スタイルの定義
    # -----------------------------
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_text = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    # -----------------------------
    # ヘッダ行の装飾
    # -----------------------------
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = thin_border
        
    # -----------------------------
    # 列幅の調整
    # -----------------------------
    column_widths = {
        "A": 10,  # カテゴリ
        "B": 12,  # 部品番号
        "C": 25,  # 品名
        "D": 8,   # 数量
        "E": 30,  # 用途 / 役割
        "F": 20,  # 材質 / 仕様
        "G": 25,  # 推奨型番
        "H": 30   # 備考 / リンク
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # -----------------------------
    # データの追加と装飾
    # -----------------------------
    for item in bom_items:
        row_data = [
            item.get("category", ""),
            item.get("part_number", ""),
            item.get("name", ""),
            item.get("quantity", 1),
            item.get("purpose", ""),
            item.get("material_spec", ""),
            item.get("supplier_pn", ""),
            item.get("remarks", "")
        ]
        ws.append(row_data)
        
        row_idx = ws.max_row
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.border = thin_border
            if col_idx in (1, 2, 4): # カテゴリ, ID, 数量は中央揃え
                cell.alignment = center_aligned_text
            else:
                cell.alignment = wrap_text

    # -----------------------------
    # カテゴリごとの背景色ルール
    # -----------------------------
    cat_a_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # 薄いオレンジ (A: オリジナル)
    cat_b_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # 薄い緑 (B: 標準品)
    cat_c_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 薄い黄色 (C: 加工素材)
    
    for row in range(2, ws.max_row + 1):
        cat_cell = ws[f"A{row}"]
        if cat_cell.value == "A":
            cat_cell.fill = cat_a_fill
        elif cat_cell.value == "B":
            cat_cell.fill = cat_b_fill
        elif cat_cell.value == "C":
            cat_cell.fill = cat_c_fill

    # -----------------------------
    # 仕上げ (ウィンドウ枠の固定と保存)
    # -----------------------------
    ws.freeze_panes = "A2"
    
    out_dir = os.path.dirname(os.path.abspath(out_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    
    wb.save(out_path)
    print(f"BOM has been successfully exported to: {out_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export JSON BOM to formatted Excel (XLSX).")
    parser.add_argument("--json", required=True, help="Input checkpoint JSON file path")
    parser.add_argument("--out", required=True, help="Output XLSX file path")
    args = parser.parse_args()
    
    export_xlsx(args.json, args.out)
